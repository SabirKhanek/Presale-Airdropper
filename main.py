import json
from openpyxl import load_workbook, Workbook
from web3 import Web3

# Configuration
sender_address = "address sending tokens"  # like : 0xD072058680B8e7Bf7c0F456DD83324D32548d7b3 
private_key = "Private key of sender address" #
contract_address = "contract_address of the presale token" # like: 0xAe0099081D9fa88DFc76cFFF903fff42a9056f80
presale_rate = 1000  # tokens per bnb
bsc_test = 'https://bsc-dataseed.binance.org/'
web = Web3(Web3.HTTPProvider(bsc_test))
if web.isConnected():
    print("Connected")

with open("abi.json") as json_file:
    abi = json.load(json_file)
contract = web.eth.contract(address=web.toChecksumAddress(contract_address), abi=abi)
decimals = contract.functions.decimals().call()

workbook = load_workbook(filename="presale_wallets.xlsx")
sheet = workbook["Wallets"]

presale_addresses = []
address_info = {}

starting_row = 2
closing_row = 44

consumed_gas = 0


def amountToDec(_amount):
    return int(_amount * pow(10, decimals))


def decToAmount(_amount):
    return _amount / pow(10, decimals)


def getBalance(_address):
    return decToAmount(contract.functions.balanceOf(_address).call())


def sendTokens(address, tokenamount):
    try:
        txn = {
            "nonce": web.eth.getTransactionCount(sender_address),
            "from": sender_address
        }
        transaction = contract.functions.transfer(address, tokenamount).buildTransaction(txn)
        transaction["gas"] += 10000
        transaction["from"] = sender_address
        signed_transaction = web.eth.account.sign_transaction(transaction, private_key)
        gasprice = (int(transaction["gas"]) * int(transaction["gasPrice"])) / pow(10, 18)
        confirmation = ""
        participated_amount = address_info[address]["participated_amount"]
        while confirmation != "confirm":
            confirmation = input(f"Transaction to {address} for the token amount {decToAmount(tokenamount)} with "
                                 f"{gasprice} BNB txn fee for {participated_amount} BNB... Write Confirm to continue: ")
            confirmation = confirmation.lower()
        global consumed_gas
        consumed_gas += gasprice
        tx_hex = web.eth.sendRawTransaction(signed_transaction.rawTransaction)
        tx_hash = web.toHex(tx_hex)
        web.eth.wait_for_transaction_receipt(tx_hash)
        explorer_url = "https://bscscan.com/tx/" + tx_hash
        # adding txn info in database
        address_info[address]["tx_hash"] = tx_hash
        address_info[address]["explorer_url"] = explorer_url
    except Exception as e:
        print(e)
        raise


def startupchecks():
    value = 0
    flag = False
    for i in range(starting_row, closing_row):
        val = sheet.cell(row=i, column=4).value
        value += val

    value = value * presale_rate

    if getBalance(sender_address) < value:
        print(f'-> Required tokens in the wallet are {value} whereas balance of token in'
              f' current account is {value - getBalance(sender_address)} less than required tokens...')
        flag = True

    recommended_gas = ((closing_row - starting_row) * 60000 * web.eth.gas_price)/pow(10, 18)
    wallet_balance = (web.eth.getBalance(sender_address))/pow(10, 18)

    if wallet_balance < recommended_gas:
        print(f'-> Transactions require around {recommended_gas},'
              f' whereas, current BNB balance is {wallet_balance},'
              f' {recommended_gas - wallet_balance} less than required balance...')

        flag = True

    if not flag:
        print(
            f"Recommended BNB balance for the transactions is: {recommended_gas} BNB and your balance is {wallet_balance} BNB")
        print(
            f"Recommended Token balance for the transactions is: {value} NUKES and your balance is {getBalance(sender_address)}  NUKES")

    return flag


if startupchecks():
    print("")
    choice = input("There are some warnings, Do you want to ignore these warnings?(Y/N)")
    if choice.lower() != 'y':
        exit()

# Preparing the Presale Wallet list
for row in range(starting_row, closing_row):
    wallet = web.toChecksumAddress(sheet.cell(row=row, column=2).value)
    if wallet not in presale_addresses:
        sent_bnb = sheet.cell(row=row, column=4).value
        if (sent_bnb >= 0.25 or sent_bnb <= 2) and (sent_bnb % 0.25) == 0.0:
            address_info[wallet] = {}
            address_info[wallet]["participated_amount"] = sent_bnb
            address_info[wallet]["exceeded_bnbs"] = 0
            presale_addresses.append(wallet)
    else:
        sent_bnb = sheet.cell(row=row, column=4).value
        current_bnb = address_info[wallet]["participated_amount"]
        if sent_bnb + current_bnb <= 2:
            address_info[wallet]["participated_amount"] = sent_bnb + current_bnb
        elif sent_bnb + current_bnb > 2:
            address_info[wallet]["participated_amount"] = 2
            try:
                address_info[wallet]["exceeded_bnbs"] += (sent_bnb + current_bnb) - 2
            except:
                address_info[wallet]["exceeded_bnbs"] = 0
                address_info[wallet]["exceeded_bnbs"] += (sent_bnb + current_bnb) - 2

# Calculating Tokens according to sent amount
for address in presale_addresses:
    address_info[address]["expected_tokens"] = address_info[address]["participated_amount"] * presale_rate

# Sending Tokens to Addresses
for address in presale_addresses:
    amount = address_info[address]["expected_tokens"]
    try:
        sendTokens(web.toChecksumAddress(address), amountToDec(amount))
        print(f'{amount} Tokens sent to {address}')
        address_info[address]["status"] = "Sent"
    except:
        print(f'Transaction failed {amount} to {address}')
        address_info[address]["status"] = "Failed"

print(f"Tokens are sent with approximate cumulative gas fees of {consumed_gas} BNB...")
print("Now calculating balances")

# Calculating Current Balances
for address in presale_addresses:
    current_balance = getBalance(address)
    address_info[address]["current_balance"] = current_balance
    if current_balance < address_info[address]["expected_tokens"]:
        address_info[address]["status"] = "Failed"

print("Saving status in spreadsheet")
# Saving status in sheet
result = Workbook()
sheet = result.active

sheet["A1"] = "Wallet Address"
sheet["B1"] = "BNB participation"
sheet["C1"] = "Expected Tokens"
sheet["D1"] = "Extra BNBs"
sheet["E1"] = "Current Balance"
sheet["F1"] = "Status"
sheet["G1"] = "Transaction Hash"
sheet["H1"] = "BSCSCAN"

row = 2
for address in address_info:
    participation = address_info[address]["participated_amount"]
    exp_token = address_info[address]["expected_tokens"]
    exceeded_bnbs = address_info[address]["exceeded_bnbs"]
    status = address_info[address]["status"]
    current_balance = address_info[address]["current_balance"]
    if status == "Sent":
        txn_hash = address_info[address]["tx_hash"]
        bscscan = address_info[address]["explorer_url"]
    elif status == "Failed":
        txn_hash = ""
        bscscan = ""
    else:
        txn_hash = ""
        bscscan = ""

    sheet["A" + str(row)] = address
    sheet["B" + str(row)] = participation
    sheet["C" + str(row)] = exp_token
    sheet["D" + str(row)] = exceeded_bnbs
    sheet["E" + str(row)] = current_balance
    sheet["F" + str(row)] = status
    sheet["G" + str(row)] = txn_hash
    sheet["H" + str(row)] = bscscan

    row += 1

result.save(filename="result.xlsx")
